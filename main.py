import requests as requests
import datetime
import re
from openpyxl import Workbook, load_workbook
import time


def add_to_excel(date, flight_no, STA, ETA, arrival_from):
    wb = load_workbook('test.xlsx')
    ws = wb.active

    max_row = ws.max_row
    max_col = ws.max_column

    print("max row= ", max_row)
    print("max col= ", max_col)

    flights_info = []
    last_ten_flights = []
    current_flight = [date, flight_no, STA, ETA]

    # print("current flight : ", current_flight)

    # checking last 10 flights, if already added then it will not be added in excel
    if max_row > 100:
        start = max_row - 100
    else:
        start = 1

    for row in range(start, max_row + 1):
        for col in range(1, max_col):
            flights_info.append(ws.cell(row=row, column=col).value)
        last_ten_flights.append(flights_info)
        flights_info = []

    # print("Last ten flights: ", last_ten_flights)

    flight_already_added = current_flight in last_ten_flights
    # print("Current flight in last ten flights: ", flight_already_added)

    if not flight_already_added:
        ws.append([date, flight_no, STA, ETA, arrival_from])
        wb.save('test.xlsx')


def epoch_to_humanreadable(epoch_time):
    return datetime.datetime.fromtimestamp(epoch_time)


epoch_time_now = time.time()  # converts time to epoch time.

while True:
    time.sleep(360)  # time in seconds
    # url = f"https://api.flightradar24.com/common/v1/airport.json?code=cjb&plugin[]=&plugin-setting[schedule][mode]=&plugin-setting[schedule][timestamp]={epoch_time_now}&page=1&limit=100&fleet=&token="

    url = f"https://api.flightradar24.com/common/v1/airport.json?code=del&plugin[]=&plugin-setting[schedule][mode]=&plugin-setting[schedule][timestamp]={epoch_time_now}&page=1&limit=100&fleet=&token="
    api_result = requests.get(url, headers={'User-Agent': 'Mozilla/5.0'})

    api_response = api_result.json()
    all_arrival_flight_data = api_response['result']['response']['airport']['pluginData']['schedule']['arrivals'][
        'data']
    for flight_data in all_arrival_flight_data:
        live_status = flight_data['flight']['status']['live']
        estimated_time_raw = flight_data['flight']['status']['text']
        estimated_time = estimated_time_raw[10:]
        actual_time_epoch = flight_data['flight']['time']['scheduled']['arrival']
        actual_time = epoch_to_humanreadable(actual_time_epoch)
        actual_time = actual_time.strftime("%H:%M")
        flight_no = flight_data['flight']['identification']['number']['default']
        arrival_from = flight_data['flight']['airport']['origin']['name']

        if live_status == True or estimated_time_raw != "Scheduled":
            date_epoch = flight_data['flight']['time']['scheduled']['arrival']
            date_human_readable = epoch_to_humanreadable(date_epoch)
            date = str(date_human_readable)[:11]
            print("Date = ", date)
            print("Flight = ", flight_no)
            print("Arrival From = ", arrival_from)
            print("Standard Arrival Time = ", actual_time)

            if estimated_time_raw == 'Canceled':
                estimated_time = 'Canceled'
            elif re.findall("^Estimated", estimated_time_raw):
                estimated_time = estimated_time_raw[10:]
                #print("Estimated Arrival Time = ", estimated_time)
            elif re.findall("^Landed", estimated_time_raw):
                estimated_time = estimated_time_raw[7:]
                print("Actual Arrival Time = ", estimated_time)
                add_to_excel(date, flight_no, actual_time, estimated_time, arrival_from)
            print("\n")
